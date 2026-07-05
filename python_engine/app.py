from __future__ import annotations

import calendar
import io
import os
from datetime import date as date_type
from datetime import datetime, timedelta, timezone
from pathlib import Path
from uuid import uuid4

from flask import Flask, abort, jsonify, render_template, request, send_file

from astronomy import get_sunrise, get_sunset, jd_to_zoned_datetime, local_date_anchor_jd
from choghadiya_service import calculate_choghadiya_slots
from yoga_service import detect_all_yogas_for_day
from location_service import geocode_city, get_timezone_name, search_locations
from pdf_generation_service import generate_pdf_export
from panchang import get_vara_from_date
from panchang_service import generate_location_panchang, resolve_location
from range_generation_service import generate_year_range_exports
from request_parsing import (
    parse_panchang_request,
    parse_pdf_generation_request,
    parse_range_generation_request,
    parse_jain_festivals_request,
)

_DAY_CHOGHADIYA_ORDER = ["Udveg", "Char", "Labh", "Amrit", "Kaal", "Shubh", "Rog"]
_NIGHT_CHOGHADIYA_ORDER = ["Shubh", "Amrit", "Char", "Rog", "Kaal", "Labh", "Udveg"]
_CHOGHADIYA_MEANINGS = {
    "Udveg": "Tension", "Amrit": "Nectar", "Rog": "Illness",
    "Labh": "Gain", "Shubh": "Auspicious", "Char": "Movement", "Kaal": "Loss",
}
_CHOGHADIYA_NATURE = {
    "Udveg": "inauspicious", "Amrit": "auspicious", "Rog": "inauspicious",
    "Labh": "auspicious", "Shubh": "auspicious", "Char": "neutral", "Kaal": "inauspicious",
}
# Vara index: Sun=0 ... Sat=6 -> starting index in the day/night Choghadiya order
_DAY_START_IDX = [0, 3, 6, 2, 5, 1, 4]
_NIGHT_START_IDX = [0, 2, 4, 6, 5, 3, 1]

GENERATED_EXPORTS: dict[str, str] = {}

_FILL_HIGHLY_AUSPICIOUS = None
_FILL_AUSPICIOUS = None
_FILL_CAUTION = None
_FILL_AVOID = None


def _get_fills():
    from openpyxl.styles import PatternFill
    return {
        "highly_auspicious": PatternFill("solid", fgColor="C6EFCE"),
        "auspicious": PatternFill("solid", fgColor="DDEBF7"),
        "caution": PatternFill("solid", fgColor="FFEB9C"),
        "avoid": PatternFill("solid", fgColor="FFC7CE"),
        "neutral": PatternFill("solid", fgColor="F2F2F2"),
    }


def _build_muhurta_workbook(summary_rows: list[dict], match_rows: list[dict]):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fills = _get_fills()
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    summary_headers = ["Date", "Vara", "Tithi", "Nakshatra", "Recommendation", "Active Yoga Count", "Active Yogas"]
    ws_sum.append(summary_headers)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws_sum[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = f"A1:{get_column_letter(len(summary_headers))}1"

    for row in summary_rows:
        ws_sum.append([
            row["date"], row["vara"], row["tithi"], row["nakshatra"],
            row["recommendation"], row["active_yoga_count"],
            row.get("active_yoga_names", ""),
        ])
        last = ws_sum.max_row
        fill = fills.get(row["recommendation"], fills["neutral"])
        for col in range(1, len(summary_headers) + 1):
            ws_sum.cell(last, col).fill = fill
            ws_sum.cell(last, col).alignment = Alignment(horizontal="center")

    for col, width in zip(range(1, 8), [14, 12, 8, 20, 20, 10, 50]):
        ws_sum.column_dimensions[get_column_letter(col)].width = width

    for row_cells in ws_sum.iter_rows(min_row=2):
        row_cells[6].alignment = Alignment(wrap_text=True)

    # ── Matches sheet ──────────────────────────────────────────────────────
    ws_match = wb.create_sheet("Matches")
    match_headers = ["Date", "Yoga Name", "Nature", "Severity", "Trigger", "Start Time", "End Time", "Meaning", "Day Recommendation"]
    ws_match.append(match_headers)

    for cell in ws_match[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    ws_match.freeze_panes = "A2"
    ws_match.auto_filter.ref = f"A1:{get_column_letter(len(match_headers))}1"

    for row in match_rows:
        ws_match.append([
            row["date"], row["yoga_name"], row["nature"], row["severity"],
            row["trigger_kind"],
            row.get("start_time", ""), row.get("end_time", ""),
            row["meaning"], row["recommendation"],
        ])
        last = ws_match.max_row
        fill = fills.get(
            "highly_auspicious" if row["nature"] == "shubh" and row["severity"] == "highly_auspicious"
            else "auspicious" if row["nature"] == "shubh"
            else "avoid" if row["severity"] == "highly_inauspicious"
            else "caution"
        )
        for col in range(1, len(match_headers) + 1):
            c = ws_match.cell(last, col)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True)

    for col, width in zip(range(1, 10), [14, 25, 12, 20, 10, 10, 10, 60, 20]):
        ws_match.column_dimensions[get_column_letter(col)].width = width

    return wb


def _add_aanandadi_sheet(wb, rows: list[dict]) -> None:
    """Append an 'Aanandadi' sheet to an existing openpyxl workbook."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fills = _get_fills()
    ws = wb.create_sheet("Aanandadi")
    headers = [
        "Date", "Yoga", "Planet", "Nakshatra", "Nature", "Severity", "Fal",
        "Start", "End", "Varjya (min)", "Varjya Start", "Varjya End",
        "Meaning", "Day Recommendation",
    ]
    ws.append(headers)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row in rows:
        varjya_val = row.get("varjya_minutes", "")
        if varjya_val == "full_day":
            varjya_display = "Full day avoid"
        elif isinstance(varjya_val, (int, float)):
            varjya_display = f"{varjya_val:.1f}"
        else:
            varjya_display = ""

        ws.append([
            row["date"], row["yoga_name"], row["planet"], row["nakshatra"],
            row["nature"], row["severity"], row["fal"],
            row.get("start_time", ""), row.get("end_time", ""),
            varjya_display,
            row.get("varjya_start", ""), row.get("varjya_end", ""),
            row["meaning"], row["recommendation"],
        ])
        last = ws.max_row
        nature = row.get("nature", "")
        severity = row.get("severity", "")
        if nature == "shubh" and severity == "highly_auspicious":
            fill = fills["highly_auspicious"]
        elif nature == "shubh":
            fill = fills["auspicious"]
        elif severity == "highly_inauspicious":
            fill = fills["avoid"]
        else:
            fill = fills["caution"]
        for col in range(1, len(headers) + 1):
            c = ws.cell(last, col)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True)

    widths = [14, 20, 12, 20, 10, 20, 10, 8, 8, 12, 12, 12, 60, 20]
    for col, width in zip(range(1, len(headers) + 1), widths):
        ws.column_dimensions[get_column_letter(col)].width = width


def _add_special_yogas_sheet(wb, rows: list[dict]) -> None:
    """Append a 'Special Yogas' sheet (Gandmool, Panchak, Jwalamukhi) to the workbook."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fills = _get_fills()
    ws = wb.create_sheet("Special Yogas")
    headers = ["Date", "Yoga", "Nature", "Severity", "Start", "End", "Trigger", "Meaning"]
    ws.append(headers)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row in rows:
        ws.append([
            row["date"], row["yoga_name"], row["nature"], row["severity"],
            row.get("start_time", ""), row.get("end_time", ""),
            row.get("trigger_detail", ""), row["meaning"],
        ])
        last = ws.max_row
        fill = fills["caution"]  # all special yogas are inauspicious
        for col in range(1, len(headers) + 1):
            ws.cell(last, col).fill = fill
            ws.cell(last, col).alignment = Alignment(wrap_text=True)

    widths = [14, 22, 10, 15, 8, 8, 30, 60]
    for col, width in zip(range(1, len(headers) + 1), widths):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_app() -> Flask:
    app = Flask(__name__)

    @app.get("/")
    def index():
        return render_template("index.html")

    @app.get("/search-location")
    def search_location():
        query = request.args.get("q", "").strip()
        if not query:
            return jsonify({"results": []})

        try:
            return jsonify({"results": search_locations(query)})
        except Exception as exc:
            return jsonify({"error": str(exc)}), 502

    @app.get("/get-coordinates")
    def get_coordinates():
        city = request.args.get("city", "").strip()
        if not city:
            return jsonify({"error": "Missing required query parameter: city"}), 400

        try:
            result = geocode_city(city)
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
        except Exception as exc:
            return jsonify({"error": str(exc)}), 502

    @app.post("/generate-panchang")
    def generate_panchang():
        try:
            body = request.get_json(silent=True) or {}
            profile = body.get("profile")
            parsed = parse_panchang_request(body)
            result = generate_location_panchang(
                parsed.input_date,
                city=parsed.city,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa_name=parsed.ayanamsa_name,
            )
            if profile:
                from jain_festival_service import generate_jain_festivals
                d_obj = datetime.strptime(parsed.input_date, "%Y-%m-%d").date()
                fest_data = generate_jain_festivals(d_obj.year, result["lat"], result["lon"], parsed.ayanamsa_name, profile)
                day_festivals = []
                for f in fest_data.get("festivals", []):
                    start_d = datetime.strptime(f["start_date"], "%Y-%m-%d").date()
                    end_d = datetime.strptime(f["end_date"], "%Y-%m-%d").date()
                    if start_d <= d_obj <= end_d:
                        day_festivals.append({
                            "occurrence_id": f["occurrence_id"],
                            "name": f["name"],
                            "category": f["category"],
                            "status": f["status"]
                        })
                result["panchang"]["jain_festivals"] = day_festivals
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/generate-range-panchang")
    def generate_range_panchang():
        try:
            parsed = parse_range_generation_request(request.get_json(silent=True))
            result = generate_year_range_exports(
                start_year=parsed.start_year,
                end_year=parsed.end_year,
                city=parsed.city,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa_name=parsed.ayanamsa_name,
                output_format=parsed.output_format,
                monthly=parsed.monthly,
                workers=parsed.workers,
                profile=parsed.profile,
            )

            files = []
            for generated_file in result["files"]:
                token = uuid4().hex
                GENERATED_EXPORTS[token] = generated_file.path
                files.append(
                    {
                        "name": generated_file.name,
                        "download_url": f"/downloads/{token}",
                    }
                )

            return jsonify(
                {
                    **{k: v for k, v in result.items() if k != "files"},
                    "files": files,
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/generate-pdf-panchang")
    def generate_pdf_panchang():
        try:
            parsed = parse_pdf_generation_request(request.get_json(silent=True))
            result = generate_pdf_export(
                year=parsed.year,
                city=parsed.city,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa_name=parsed.ayanamsa_name,
            )
            token = uuid4().hex
            GENERATED_EXPORTS[token] = result["file"]["path"]

            return jsonify(
                {
                    "year": result["year"],
                    "ayanamsa": result["ayanamsa"],
                    "location": result["location"],
                    "file": {
                        "name": result["file"]["name"],
                        "download_url": f"/downloads/{token}",
                    },
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.get("/month-overview")
    def month_overview():
        try:
            year_str = request.args.get("year")
            month_str = request.args.get("month")

            if not year_str:
                return jsonify({"error": "Missing required parameter: year"}), 400
            if not month_str:
                return jsonify({"error": "Missing required parameter: month"}), 400

            try:
                year = int(year_str)
                month = int(month_str)
            except ValueError:
                return jsonify({"error": "year and month must be integers"}), 400

            if not (1 <= month <= 12):
                return jsonify({"error": "month must be between 1 and 12"}), 400
            if not (1900 <= year <= 2200):
                return jsonify({"error": "year must be between 1900 and 2200"}), 400

            city = request.args.get("city") or None
            lat_str = request.args.get("lat")
            lon_str = request.args.get("lon")
            ayanamsa = request.args.get("ayanamsa", "Lahiri")

            lat = float(lat_str) if lat_str else None
            lon = float(lon_str) if lon_str else None

            if not city and not (lat is not None and lon is not None):
                return jsonify({"error": "Provide either a city name or both latitude and longitude."}), 400

            location = resolve_location(city=city, lat=lat, lon=lon)

            num_days = calendar.monthrange(year, month)[1]
            days = []
            first_result = None

            profile = request.args.get("profile")
            date_to_festivals = {}
            if profile:
                from jain_festival_service import generate_jain_festivals
                fest_data = generate_jain_festivals(year, location.lat, location.lon, ayanamsa, profile)
                for f in fest_data.get("festivals", []):
                    start_d = datetime.strptime(f["start_date"], "%Y-%m-%d").date()
                    end_d = datetime.strptime(f["end_date"], "%Y-%m-%d").date()
                    curr_d = start_d
                    while curr_d <= end_d:
                        if curr_d.year == year and curr_d.month == month:
                            d_str = curr_d.isoformat()
                            if d_str not in date_to_festivals:
                                date_to_festivals[d_str] = []
                            if not any(x["occurrence_id"] == f["occurrence_id"] for x in date_to_festivals[d_str]):
                                date_to_festivals[d_str].append({
                                    "occurrence_id": f["occurrence_id"],
                                    "name": f["name"],
                                    "category": f["category"],
                                    "status": f["status"]
                                })
                        curr_d += timedelta(days=1)

            for day_num in range(1, num_days + 1):
                date_str = f"{year:04d}-{month:02d}-{day_num:02d}"
                result = generate_location_panchang(
                    date_str,
                    lat=location.lat,
                    lon=location.lon,
                    ayanamsa_name=ayanamsa,
                )
                if first_result is None:
                    first_result = result

                udaya_tithi = result["panchang"]["tithi"][0]
                tithi_index = udaya_tithi["index"]
                nakshatra_index = result["panchang"]["nakshatra"]["index"]
                vara_index = result["panchang"]["vara"]["index"]

                tithi_end_raw = udaya_tithi["ends"]["time"] if udaya_tithi["ends"] else ""
                nakshatra_end_raw = result["panchang"]["nakshatra"]["ends"]["time"]
                
                day_payload = {
                    "date": date_str,
                    "tithi_index": tithi_index,
                    "tithi_name": udaya_tithi["name"],
                    "tithi_end_time": tithi_end_raw[:5] if tithi_end_raw else "",
                    "nakshatra_index": nakshatra_index,
                    "nakshatra_name": result["panchang"]["nakshatra"]["name"],
                    "nakshatra_end_time": nakshatra_end_raw[:5] if nakshatra_end_raw else "",
                    "vara_index": vara_index,
                    "vara_name": result["panchang"]["vara"]["name"],
                    "is_purnima": tithi_index == 15,
                    "is_amavasya": tithi_index == 30,
                    "is_ekadashi": tithi_index in (11, 26),
                    "sunrise_time": result["events"]["sunrise"]["time"][:5]
                        if result.get("events", {}).get("sunrise", {}).get("time") else "",
                    "sunset_time": result["events"]["sunset"]["time"][:5]
                        if result.get("events", {}).get("sunset", {}).get("time") else "",
                    "has_panchak": result.get("panchak_kaal", {}).get("has_window", False),
                }
                if profile:
                    day_payload["jain_festivals"] = date_to_festivals.get(date_str, [])
                days.append(day_payload)

            hindu_month_index = first_result["panchang"]["hindu_month"]["index"] if first_result else 0
            hindu_month = first_result["panchang"]["hindu_month"]["name"] if first_result else ""
            vikram_samvat = first_result["panchang"]["vikram_samvat"] if first_result else year + 57

            return jsonify({
                "year": year,
                "month": month,
                "location": location.name,
                "timezone": location.timezone,
                "hindu_month": hindu_month,
                "hindu_month_index": hindu_month_index,
                "vikram_samvat": vikram_samvat,
                "days": days,
            })
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/generate-jain-festivals")
    def generate_jain_festivals_api():
        try:
            body = request.get_json(silent=True) or {}
            parsed = parse_jain_festivals_request(body)
            from jain_festival_service import generate_jain_festivals
            result = generate_jain_festivals(
                year=parsed.year,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa=parsed.ayanamsa_name,
                profile=parsed.profile
            )
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/generate-jain-festival-exports")
    def generate_jain_festival_exports_api():
        try:
            body = request.get_json(silent=True) or {}
            parsed = parse_jain_festivals_request(body)
            fmt = body.get("format", "csv").lower()
            if fmt not in {"csv", "excel", "json", "pdf", "all"}:
                return jsonify({"error": "format must be one of: csv, excel, json, pdf, all"}), 400
                
            from jain_festival_service import generate_jain_festivals
            result = generate_jain_festivals(
                year=parsed.year,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa=parsed.ayanamsa_name,
                profile=parsed.profile
            )
            
            flat_rows = []
            for f in result.get("festivals", []):
                flat_rows.append({
                    "Festival_ID": f["id"],
                    "Name_English": f["name"],
                    "Name_Hindi": f["name_hindi"],
                    "Category": f["category"],
                    "Start_Date": f["start_date"],
                    "End_Date": f["end_date"],
                    "Jain_Month": f["jain_month"],
                    "Paksha": f["paksha"],
                    "Tithi": f["tithi"],
                    "Profile": f["profile"],
                    "Status": f["status"],
                    "Meaning": f["meaning"],
                    "Observance": f["observance"],
                    "Sources": "; ".join(f["sources"])
                })
                
            import tempfile
            import os
            from uuid import uuid4
            
            tmpdir = tempfile.gettempdir()
            unique_id = uuid4().hex
            files = []
            
            formats_to_gen = ["csv", "json", "excel", "pdf"] if fmt == "all" else [fmt]
            for f_format in formats_to_gen:
                if f_format == "csv":
                    filename = f"jain_festivals_{parsed.year}_{unique_id}.csv"
                    path = os.path.join(tmpdir, filename)
                    import csv
                    with open(path, 'w', newline='', encoding='utf-8-sig') as f_out:
                        writer = csv.DictWriter(f_out, fieldnames=list(flat_rows[0].keys()) if flat_rows else [])
                        writer.writeheader()
                        writer.writerows(flat_rows)
                    token = uuid4().hex
                    GENERATED_EXPORTS[token] = path
                    files.append({
                        "name": f"jain_festivals_{parsed.year}.csv",
                        "download_url": f"/downloads/{token}"
                    })
                elif f_format == "json":
                    filename = f"jain_festivals_{parsed.year}_{unique_id}.json"
                    path = os.path.join(tmpdir, filename)
                    import json
                    with open(path, 'w', encoding='utf-8') as f_out:
                        json.dump(flat_rows, f_out, indent=2, ensure_ascii=False)
                    token = uuid4().hex
                    GENERATED_EXPORTS[token] = path
                    files.append({
                        "name": f"jain_festivals_{parsed.year}.json",
                        "download_url": f"/downloads/{token}"
                    })
                elif f_format == "excel":
                    filename = f"jain_festivals_{parsed.year}_{unique_id}.xlsx"
                    path = os.path.join(tmpdir, filename)
                    import pandas as pd
                    df = pd.DataFrame(flat_rows)
                    df.to_excel(path, index=False, engine='openpyxl')
                    token = uuid4().hex
                    GENERATED_EXPORTS[token] = path
                    files.append({
                        "name": f"jain_festivals_{parsed.year}.xlsx",
                        "download_url": f"/downloads/{token}"
                    })
                elif f_format == "pdf":
                    filename = f"jain_festivals_{parsed.year}_{unique_id}.pdf"
                    path = os.path.join(tmpdir, filename)
                    
                    from reportlab.lib.pagesizes import letter
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib import colors
                    
                    doc = SimpleDocTemplate(path, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
                    styles = getSampleStyleSheet()
                    
                    title_style = ParagraphStyle(
                        'TitleStyle',
                        parent=styles['Heading1'],
                        fontSize=18,
                        textColor=colors.HexColor('#2C3E50'),
                        spaceAfter=12
                    )
                    
                    story = []
                    story.append(Paragraph(f"Jain Festival Calendar - {parsed.year}", title_style))
                    story.append(Paragraph(f"Profile: {parsed.profile.replace('_', ' ').title()}", styles['Normal']))
                    story.append(Paragraph(f"Location: Lat {parsed.lat}, Lon {parsed.lon} ({result['location']['timezone']})", styles['Normal']))
                    story.append(Spacer(1, 12))
                    
                    headers = ["Date", "Name (English)", "Category", "Jain Month", "Tithi", "Observance"]
                    table_data = [headers]
                    
                    for row in flat_rows:
                        obs_para = Paragraph(row["Observance"], styles["BodyText"])
                        table_data.append([
                            row["Start_Date"],
                            row["Name_English"],
                            row["Category"].title(),
                            row["Jain_Month"],
                            str(row["Tithi"]),
                            obs_para
                        ])
                        
                    t = Table(table_data, colWidths=[70, 110, 60, 70, 40, 190])
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C3E50')),
                        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('VALIGN', (0,0), (-1,-1), 'TOP'),
                        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0,0), (-1,0), 10),
                        ('BOTTOMPADDING', (0,0), (-1,0), 6),
                        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F8F9FA')),
                        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F4F4')]),
                        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D0D3D4')),
                        ('FONTSIZE', (0,1), (-1,-1), 9),
                    ]))
                    
                    story.append(t)
                    doc.build(story)
                    
                    token = uuid4().hex
                    GENERATED_EXPORTS[token] = path
                    files.append({
                        "name": f"jain_festivals_{parsed.year}.pdf",
                        "download_url": f"/downloads/{token}"
                    })
                    
            return jsonify({
                "year": parsed.year,
                "profile": parsed.profile,
                "files": files
            })
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/choghadiya")
    def choghadiya():
        try:
            body = request.get_json(silent=True) or {}
            date_str = body.get("date")
            lat = body.get("lat")
            lon = body.get("lon")

            if not date_str:
                return jsonify({"error": "Missing required field: date"}), 400
            if lat is None or lon is None:
                return jsonify({"error": "Missing required fields: lat and lon"}), 400

            try:
                parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"error": "date must be in YYYY-MM-DD format"}), 400

            tz_name = get_timezone_name(float(lat), float(lon))
            anchor_jd = local_date_anchor_jd(parsed_date, tz_name)
            sunrise_jd = get_sunrise(anchor_jd, float(lat), float(lon))
            sunset_jd = get_sunset(anchor_jd, float(lat), float(lon))

            next_date = parsed_date + timedelta(days=1)
            next_anchor_jd = local_date_anchor_jd(next_date, tz_name)
            next_sunrise_jd = get_sunrise(next_anchor_jd, float(lat), float(lon))

            if not sunrise_jd or not sunset_jd or not next_sunrise_jd:
                return jsonify({"error": "Sunrise or sunset could not be calculated for this location/date."}), 400

            weekday = (parsed_date.weekday() + 1) % 7
            sunrise_dt = jd_to_zoned_datetime(sunrise_jd, tz_name)
            sunset_dt = jd_to_zoned_datetime(sunset_jd, tz_name)
            sunrise_utc = sunrise_dt.astimezone(timezone.utc) if sunrise_dt else None
            sunset_utc = sunset_dt.astimezone(timezone.utc) if sunset_dt else None
            day_slot_duration_minutes = (sunset_jd - sunrise_jd) * 1440 / 8
            night_slot_duration_minutes = (next_sunrise_jd - sunset_jd) * 1440 / 8

            slots = calculate_choghadiya_slots(
                sunrise_jd, sunset_jd, next_sunrise_jd, weekday, tz_name, parsed_date
            )

            return jsonify({
                "date": date_str,
                "timezone": tz_name,
                "sunrise": sunrise_dt.strftime("%H:%M") if sunrise_dt else "",
                "sunset": sunset_dt.strftime("%H:%M") if sunset_dt else "",
                "sunrise_local": sunrise_dt.isoformat(timespec="seconds") if sunrise_dt else "",
                "sunset_local": sunset_dt.isoformat(timespec="seconds") if sunset_dt else "",
                "sunrise_utc": sunrise_utc.isoformat(timespec="seconds") if sunrise_utc else "",
                "sunset_utc": sunset_utc.isoformat(timespec="seconds") if sunset_utc else "",
                "day_slot_duration_minutes": round(day_slot_duration_minutes, 3),
                "night_slot_duration_minutes": round(night_slot_duration_minutes, 3),
                "slots": slots,
            })
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/dainika-muhurta")
    def dainika_muhurta():
        try:
            body = request.get_json(silent=True) or {}
            date_str = body.get("date")
            lat = body.get("lat")
            lon = body.get("lon")
            ayanamsa = body.get("ayanamsa", "Lahiri")

            if not date_str:
                return jsonify({"error": "Missing required field: date"}), 400
            if lat is None or lon is None:
                return jsonify({"error": "Missing required fields: lat and lon"}), 400

            try:
                parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"error": "date must be in YYYY-MM-DD format"}), 400

            tz_name = get_timezone_name(float(lat), float(lon))
            anchor_jd = local_date_anchor_jd(parsed_date, tz_name)
            next_anchor_jd = local_date_anchor_jd(parsed_date + timedelta(days=1), tz_name)
            sunrise_jd = get_sunrise(anchor_jd, float(lat), float(lon))
            next_sunrise_jd = get_sunrise(next_anchor_jd, float(lat), float(lon))

            if not sunrise_jd or not next_sunrise_jd:
                return jsonify({"error": "Sunrise could not be calculated for this location/date."}), 400

            result = detect_all_yogas_for_day(
                date_obj=parsed_date,
                sunrise_jd=sunrise_jd,
                next_sunrise_jd=next_sunrise_jd,
                tz_name=tz_name,
                ayanamsa=ayanamsa,
            )
            return jsonify({"date": date_str, **result})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/dainika-muhurta-export")
    def dainika_muhurta_export():
        try:
            body = request.get_json(silent=True) or {}
            year = body.get("year")
            month = body.get("month")
            lat = body.get("lat")
            lon = body.get("lon")
            ayanamsa = body.get("ayanamsa", "Lahiri")

            if year is None:
                return jsonify({"error": "Missing required field: year"}), 400
            if month is None:
                return jsonify({"error": "Missing required field: month"}), 400
            if lat is None or lon is None:
                return jsonify({"error": "Missing required fields: lat and lon"}), 400

            year, month = int(year), int(month)
            if not (1 <= month <= 12):
                return jsonify({"error": "month must be 1–12"}), 400

            tz_name = get_timezone_name(float(lat), float(lon))
            days_in_month = calendar.monthrange(year, month)[1]

            summary_rows = []
            match_rows = []
            aanandadi_rows = []
            special_rows = []

            NAKSHATRA_NAMES = [
                "", "Ashvini", "Bharani", "Kritika", "Rohini", "Mrigashira",
                "Ardra", "Punarvasu", "Pushya", "Ashlesha", "Magha",
                "Purva Phalguni", "Uttara Phalguni", "Hasta", "Chitra", "Swati",
                "Vishakha", "Anuradha", "Jyeshtha", "Mula", "Purva Ashadha",
                "Uttara Ashadha", "Shravana", "Dhanishtha", "Shatabhisha",
                "Purva Bhadrapada", "Uttara Bhadrapada", "Revati", "Abhijit",
            ]
            VARA_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

            for day in range(1, days_in_month + 1):
                d = date_type(year, month, day)
                anchor_jd = local_date_anchor_jd(d, tz_name)
                next_anchor_jd = local_date_anchor_jd(
                    date_type(year, month, day + 1) if day < days_in_month
                    else date_type(year + (month == 12), (month % 12) + 1, 1),
                    tz_name,
                )
                sunrise_jd = get_sunrise(anchor_jd, float(lat), float(lon))
                next_sunrise_jd = get_sunrise(next_anchor_jd, float(lat), float(lon))
                if not sunrise_jd or not next_sunrise_jd:
                    continue

                day_result = detect_all_yogas_for_day(
                    date_obj=d,
                    sunrise_jd=sunrise_jd,
                    next_sunrise_jd=next_sunrise_jd,
                    tz_name=tz_name,
                    ayanamsa=ayanamsa,
                )

                vara = day_result["vara"]
                tithi = day_result["tithi"]
                nakshatra = day_result["nakshatra"]
                active_yogas = [y for y in day_result["yogas"] if not y.get("cancelled")]
                summary_rows.append({
                    "date": d.strftime("%Y-%m-%d"),
                    "vara": VARA_NAMES[vara],
                    "tithi": tithi,
                    "nakshatra": NAKSHATRA_NAMES[nakshatra] if nakshatra < len(NAKSHATRA_NAMES) else str(nakshatra),
                    "recommendation": day_result["recommendation"],
                    "active_yoga_count": len(active_yogas),
                    "active_yoga_names": ", ".join(y["name"] for y in active_yogas),
                })
                for yoga in day_result["yogas"]:
                    if yoga.get("cancelled"):
                        continue
                    match_rows.append({
                        "date": d.strftime("%Y-%m-%d"),
                        "yoga_name": yoga["name"],
                        "nature": yoga["nature"],
                        "severity": yoga["severity"],
                        "trigger_kind": yoga["trigger_kind"],
                        "meaning": yoga["meaning"],
                        "recommendation": day_result["recommendation"],
                        "start_time": yoga.get("start_time", ""),
                        "end_time": yoga.get("end_time", ""),
                    })
                for yoga in day_result["special_yogas"]:
                    special_rows.append({
                        "date": d.strftime("%Y-%m-%d"),
                        "yoga_name": yoga["name"],
                        "nature": yoga["nature"],
                        "severity": yoga["severity"],
                        "start_time": yoga.get("start_time", ""),
                        "end_time": yoga.get("end_time", ""),
                        "meaning": yoga["meaning"],
                        "trigger_detail": yoga.get("trigger_detail", ""),
                    })
                for yoga in day_result["aanandadi_yogas"]:
                    aanandadi_rows.append({
                        "date": d.strftime("%Y-%m-%d"),
                        "yoga_name": yoga["name"],
                        "planet": yoga["triggering_planet"],
                        "nakshatra": yoga["trigger_nakshatra"],
                        "nature": yoga["nature"],
                        "severity": yoga["severity"],
                        "fal": yoga["fal"],
                        "start_time": yoga.get("start_time", ""),
                        "end_time": yoga.get("end_time", ""),
                        "varjya_minutes": yoga.get("varjya_minutes", ""),
                        "varjya_start": yoga.get("varjya_start_time", ""),
                        "varjya_end": yoga.get("varjya_end_time", ""),
                        "meaning": yoga["meaning"],
                        "recommendation": day_result["aanandadi_recommendation"],
                    })

            wb = _build_muhurta_workbook(summary_rows, match_rows)
            _add_aanandadi_sheet(wb, aanandadi_rows)
            _add_special_yogas_sheet(wb, special_rows)
            filename = f"dainika_muhurta_{year}_{str(month).zfill(2)}.xlsx"

            tmp_dir = Path("output")
            tmp_dir.mkdir(exist_ok=True)
            file_path = tmp_dir / filename
            wb.save(str(file_path))

            token = str(uuid4())
            GENERATED_EXPORTS[token] = str(file_path)

            return jsonify({
                "filename": filename,
                "download_url": f"/downloads/{token}",
            })
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.get("/downloads/<token>")
    def download_generated_file(token: str):
        file_path = GENERATED_EXPORTS.get(token)
        if not file_path:
            abort(404)

        path = Path(file_path)
        if not path.exists():
            abort(404)

        return send_file(path, as_attachment=True, download_name=path.name)

    # ── Panchang DB generation endpoints ────────────────────────────────────

    _DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
    _generation_status: dict[str, dict] = {}

    @app.post("/api/generate-db")
    def api_generate_db():
        import threading
        from db_generator import generate_panchang_db

        body = request.get_json(silent=True) or {}
        city_name = body.get("city_name")
        city_slug = body.get("city_slug")
        latitude = body.get("latitude")
        longitude = body.get("longitude")
        tz = body.get("timezone")

        if not all([city_name, city_slug, latitude is not None, longitude is not None, tz]):
            return jsonify({"error": "Required fields: city_name, city_slug, latitude, longitude, timezone"}), 400

        db_path = os.path.join(_DATA_DIR, f"panchang_{city_slug}.db")
        _generation_status[city_slug] = {"status": "running", "progress": 0, "message": ""}

        def _run():
            try:
                def _progress(done: int, total: int) -> None:
                    pct = int(done / total * 100) if total else 0
                    _generation_status[city_slug]["progress"] = pct

                generate_panchang_db(
                    city_name=city_name,
                    city_slug=city_slug,
                    latitude=float(latitude),
                    longitude=float(longitude),
                    timezone_str=tz,
                    db_path=db_path,
                    progress_callback=_progress,
                )
                _generation_status[city_slug] = {"status": "complete", "progress": 100, "message": ""}
            except Exception as exc:
                _generation_status[city_slug] = {"status": "error", "progress": 0, "message": str(exc)}

        threading.Thread(target=_run, daemon=True).start()
        return jsonify({"status": "started", "city_slug": city_slug})

    @app.get("/api/generate-db/progress/<city_slug>")
    def api_generate_db_progress(city_slug: str):
        status = _generation_status.get(city_slug)
        if status is None:
            db_path = os.path.join(_DATA_DIR, f"panchang_{city_slug}.db")
            if os.path.isfile(db_path):
                return jsonify({"status": "complete", "progress": 100, "message": ""})
            return jsonify({"status": "not_started", "progress": 0, "message": ""}), 404
        return jsonify(status)

    @app.delete("/api/generate-db/<city_slug>")
    def api_delete_db(city_slug: str):
        db_path = os.path.join(_DATA_DIR, f"panchang_{city_slug}.db")
        _generation_status.pop(city_slug, None)
        if os.path.isfile(db_path):
            try:
                os.remove(db_path)
            except OSError as exc:
                return jsonify({"error": str(exc)}), 500
        return jsonify({"deleted": city_slug})

    return app


app = create_app()


if __name__ == "__main__":
    app.run(debug=True)
