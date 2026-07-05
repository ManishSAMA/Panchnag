import os
import sys
import math
import threading
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Adjust sys.path to find sibling modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sun_transit_exporter.calculation import generate_transit_events

# Indian Cities with coordinates and timezone details
CITIES = {
    "Delhi": {
        "coords": (77.2090, 28.6139, 216.0),  # Longitude E, Latitude N, Altitude m
        "timezone": "Asia/Kolkata"
    },
    "Mumbai": {
        "coords": (72.8777, 19.0760, 14.0),
        "timezone": "Asia/Kolkata"
    },
    "Kolkata": {
        "coords": (88.3639, 22.5726, 9.0),
        "timezone": "Asia/Kolkata"
    },
    "Chennai": {
        "coords": (80.2707, 13.0827, 6.0),
        "timezone": "Asia/Kolkata"
    }
}

class SunTransitApp(tk.Tk):
    def __init__(self):
        super().__init__()
        
        self.title("Sun Transit Excel Exporter")
        self.geometry("600x530")
        self.minsize(550, 480)
        
        # Set custom window icon
        self.configure(bg="#f8f9fa")
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "icon.ico")
        if os.path.exists(icon_path):
            try:
                self.iconbitmap(icon_path)
            except Exception:
                pass  # Fallback if OS doesn't support iconbitmap
        
        # Set style configuration
        self.style = ttk.Style(self)
        self.style.theme_use("clam")
        self.configure_styles()
        
        self.last_generated_file = None
        self.create_widgets()
        
    def configure_styles(self):
        # Base frame bg
        self.style.configure(".", background="#f8f9fa", foreground="#212529", font=("Segoe UI", 10))
        
        # Card style frame
        self.style.configure("Card.TFrame", background="#ffffff", relief="flat", borderwidth=1)
        
        # Primary action button
        self.style.configure("Primary.TButton", 
                             background="#0077b6", 
                             foreground="#ffffff", 
                             font=("Segoe UI", 10, "bold"), 
                             padding=6, 
                             relief="flat")
        self.style.map("Primary.TButton",
                       background=[("active", "#0096c7"), ("disabled", "#adb5bd")],
                       foreground=[("disabled", "#6c757d")])
                       
        # Standard buttons
        self.style.configure("Standard.TButton", 
                             background="#e9ecef", 
                             foreground="#212529", 
                             font=("Segoe UI", 10), 
                             padding=4, 
                             relief="flat")
        self.style.map("Standard.TButton",
                       background=[("active", "#dee2e6")])

        # Accent button (for open Excel)
        self.style.configure("Accent.TButton", 
                             background="#2a9d8f", 
                             foreground="#ffffff", 
                             font=("Segoe UI", 10, "bold"), 
                             padding=6, 
                             relief="flat")
        self.style.map("Accent.TButton",
                       background=[("active", "#3dbeb2"), ("disabled", "#adb5bd")],
                       foreground=[("disabled", "#6c757d")])

        # Progressbar
        self.style.configure("Custom.Horizontal.TProgressbar", 
                             troughcolor="#e9ecef", 
                             background="#0077b6", 
                             thickness=15)
                             
    def create_widgets(self):
        # 1. Header block
        header_frame = ttk.Frame(self, style="TFrame")
        header_frame.pack(fill="x", padx=20, pady=(15, 5))
        
        title_label = ttk.Label(header_frame, 
                                text="Sun Transit Excel Exporter", 
                                font=("Segoe UI", 16, "bold"), 
                                foreground="#0077b6",
                                background="#f8f9fa")
        title_label.pack(anchor="w")
        
        desc_label = ttk.Label(header_frame, 
                               text="Calculate the exact times of Sun's tropical longitude crossing every 1-arcminute boundary.", 
                               foreground="#6c757d",
                               background="#f8f9fa")
        desc_label.pack(anchor="w", pady=(2, 0))
        
        # 2. Form Container (Card)
        form_card = ttk.Frame(self, style="Card.TFrame")
        form_card.pack(fill="x", padx=20, pady=5)
        
        # Internal padding inside the card
        form_inner = tk.Frame(form_card, bg="#ffffff", bd=0)
        form_inner.pack(fill="both", padx=15, pady=15)
        
        # Years grid
        ttk.Label(form_inner, text="Start Year:", background="#ffffff").grid(row=0, column=0, sticky="w", pady=5)
        self.start_year_var = tk.StringVar(value=str(datetime.now().year))
        self.start_year_spin = ttk.Spinbox(form_inner, from_=1800, to=2200, textvariable=self.start_year_var, width=10)
        self.start_year_spin.grid(row=0, column=1, sticky="w", padx=(5, 20), pady=5)
        
        ttk.Label(form_inner, text="End Year:", background="#ffffff").grid(row=0, column=2, sticky="w", pady=5)
        self.end_year_var = tk.StringVar(value=str(datetime.now().year))
        self.end_year_spin = ttk.Spinbox(form_inner, from_=1800, to=2200, textvariable=self.end_year_var, width=10)
        self.end_year_spin.grid(row=0, column=3, sticky="w", padx=5, pady=5)
        
        # City Dropdown
        ttk.Label(form_inner, text="City (Exact Coordinates):", background="#ffffff").grid(row=1, column=0, sticky="w", pady=8)
        self.city_var = tk.StringVar(value="Delhi")
        self.city_combo = ttk.Combobox(form_inner, textvariable=self.city_var, values=list(CITIES.keys()), state="readonly", width=12)
        self.city_combo.grid(row=1, column=1, columnspan=3, sticky="w", padx=(5, 0), pady=8)
        
        # Save directory
        ttk.Label(form_inner, text="Output Directory:", background="#ffffff").grid(row=2, column=0, sticky="w", pady=5)
        self.output_dir_var = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Desktop"))
        # If Desktop doesn't exist, default to current working directory
        if not os.path.exists(self.output_dir_var.get()):
            self.output_dir_var.set(os.getcwd())
            
        self.output_dir_entry = ttk.Entry(form_inner, textvariable=self.output_dir_var, width=45)
        self.output_dir_entry.grid(row=2, column=1, columnspan=2, sticky="ew", padx=(5, 5), pady=5)
        
        self.browse_btn = ttk.Button(form_inner, text="Browse...", style="Standard.TButton", command=self.browse_directory)
        self.browse_btn.grid(row=2, column=3, sticky="e", pady=5)
        
        # Grid column weights
        form_inner.grid_columnconfigure(2, weight=1)
        
        # 3. Actions block (Placed directly under inputs for visibility!)
        actions_frame = ttk.Frame(self)
        actions_frame.pack(fill="x", padx=20, pady=10)
        
        self.generate_btn = ttk.Button(actions_frame, text="Generate Excel Report", style="Primary.TButton", command=self.start_generation)
        self.generate_btn.pack(side="left", padx=(0, 10))
        
        self.open_btn = ttk.Button(actions_frame, text="Open Excel File", style="Accent.TButton", command=self.open_generated_file, state="disabled")
        self.open_btn.pack(side="left")
        
        # 4. Progress and status container
        status_frame = ttk.Frame(self)
        status_frame.pack(fill="x", padx=20, pady=5)
        
        self.progress_bar = ttk.Progressbar(status_frame, style="Custom.Horizontal.TProgressbar", mode="determinate")
        self.progress_bar.pack(fill="x", pady=(0, 2))
        
        self.status_var = tk.StringVar(value="Ready to generate.")
        self.status_label = ttk.Label(status_frame, textvariable=self.status_var, font=("Segoe UI", 9, "italic"), foreground="#495057")
        self.status_label.pack(anchor="w")
        
        # 5. Logging Text view (Fixed height so it doesn't push anything out of view)
        log_frame = ttk.Frame(self, style="Card.TFrame")
        log_frame.pack(fill="both", expand=True, padx=20, pady=(5, 15))
        
        scrollbar = ttk.Scrollbar(log_frame)
        scrollbar.pack(side="right", fill="y")
        
        self.log_text = tk.Text(log_frame, wrap="word", yscrollcommand=scrollbar.set, font=("Consolas", 9), bg="#fafafa", fg="#333333", bd=0, height=7)
        self.log_text.pack(fill="both", expand=True, padx=5, pady=5)
        scrollbar.config(command=self.log_text.yview)
        
        # Prevent manual user editing of logs
        self.log_text.config(state="disabled")
        
    def log(self, message: str):
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")
        
    def browse_directory(self):
        dir_path = filedialog.askdirectory(initialdir=self.output_dir_var.get())
        if dir_path:
            self.output_dir_var.set(dir_path)
            
    def open_generated_file(self):
        if self.last_generated_file and os.path.exists(self.last_generated_file):
            try:
                os.startfile(self.last_generated_file)
            except Exception as e:
                messagebox.showerror("Error", f"Could not open file: {e}")
        else:
            messagebox.showerror("Error", "File does not exist or has been moved.")
            
    def start_generation(self):
        # Validate years
        try:
            start_year = int(self.start_year_var.get())
            end_year = int(self.end_year_var.get())
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter valid integers for the years.")
            return
            
        if start_year > end_year:
            messagebox.showerror("Invalid Range", "Start Year must be less than or equal to End Year.")
            return
            
        # Disable controls during calculation
        self.generate_btn.config(state="disabled")
        self.open_btn.config(state="disabled")
        self.start_year_spin.config(state="disabled")
        self.end_year_spin.config(state="disabled")
        self.city_combo.config(state="disabled")
        self.browse_btn.config(state="disabled")
        
        # Clear logs
        self.log_text.config(state="normal")
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state="disabled")
        
        # Reset progress
        self.progress_bar["value"] = 0
        
        # Run calculation in a background thread to keep UI responsive
        thread = threading.Thread(target=self.run_export_thread, args=(start_year, end_year))
        thread.daemon = True
        thread.start()
        
    def run_export_thread(self, start_year: int, end_year: int):
        city_name = self.city_var.get()
        city_info = CITIES[city_name]
        topo_coords = city_info["coords"]
        tz_name = city_info["timezone"]
        tz = ZoneInfo(tz_name)
        
        self.log(f"Starting Sun Transit generation for {start_year} to {end_year}...")
        self.log(f"Selected Location: {city_name} (Lat: {topo_coords[1]}°, Lon: {topo_coords[0]}°, Alt: {topo_coords[2]}m)")
        self.log(f"Local timezone context: {tz_name}")
        
        try:
            # Build list of years to process
            years = list(range(start_year, end_year + 1))
            total_years = len(years)
            
            rows = []
            
            for idx, year in enumerate(years):
                self.status_var.set(f"Generating data for year {year}...")
                self.log(f"Computing transits for year {year}...")
                
                # Start and end of the year in local timezone
                dt_start = datetime(year, 1, 1, 0, 0, 0, tzinfo=tz)
                dt_end = datetime(year, 12, 31, 23, 59, 59, tzinfo=tz)
                
                # Generate transits for this year
                year_events = list(generate_transit_events(dt_start, dt_end, topo_coords=topo_coords))
                self.log(f"Found {len(year_events):,} crossings in {year}.")
                
                # Traditional Rashi Names
                rashi_names = [
                    "Mesh", "Vrishabh", "Mithun", "Kark", "Simha", "Kanya",
                    "Tula", "Vrishchik", "Dhanu", "Makar", "Kumbh", "Meen"
                ]

                # Convert crossing UTC datetimes to local timezone for the output
                for event in year_events:
                    dt_local = event['dt'].astimezone(tz)
                    
                    deg = event['degree']
                    rashi_idx = deg // 30
                    ansha = deg % 30
                    kala = event['minute']
                    vikala = 0
                    
                    rashi_name = rashi_names[rashi_idx]
                    
                    rows.append({
                        "Date": dt_local.strftime("%Y-%m-%d"),
                        "Time": dt_local.strftime("%H:%M:%S"),
                        "Rashi": rashi_name,
                        "Ansha": ansha,
                        "Kala": kala,
                        "Vikala": vikala,
                        "Ayanamsa_DM": event['ayanamsa']
                    })
                    
                # Update progress
                progress = ((idx + 1) / total_years) * 80  # Save last 20% for formatting Excel
                self.progress_bar["value"] = progress
                
            # Create DataFrame
            self.status_var.set("Formatting data...")
            self.log(f"Collected {len(rows):,} total crossing records. Constructing data sheet...")
            df = pd.DataFrame(rows)
            
            # Format filename
            filename = f"Sun_Transits_{city_name}_{start_year}_{end_year}.xlsx"
            file_path = os.path.join(self.output_dir_var.get(), filename)
            
            self.log("Writing to Excel and styling the spreadsheet...")
            self.save_to_excel_styled(df, file_path, city_name, start_year, end_year)
            
            self.progress_bar["value"] = 100
            self.last_generated_file = file_path
            self.status_var.set("Export completed successfully!")
            self.log(f"Excel report saved successfully to:")
            self.log(file_path)
            
            # Enable Open Button
            self.open_btn.config(state="normal")
            messagebox.showinfo("Success", f"Excel report generated successfully!\nSaved to: {file_path}")
            
        except Exception as e:
            self.status_var.set("Failed!")
            self.log(f"ERROR: An error occurred during export: {e}")
            messagebox.showerror("Export Failed", f"An error occurred during generation:\n{e}")
            
        finally:
            # Re-enable UI components
            self.generate_btn.config(state="normal")
            self.start_year_spin.config(state="normal")
            self.end_year_spin.config(state="normal")
            self.city_combo.config(state="normal")
            self.browse_btn.config(state="normal")
            
    def save_to_excel_styled(self, df: pd.DataFrame, file_path: str, city_name: str, start_year: int, end_year: int):
        sheet_name = f"Transits {start_year}-{end_year}"
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # 1. Colors & Fills (Professional deep blue)
            header_fill = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Style header row (Row 1)
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
            # 2. Predefined Column dimensions (Avoids scanning 40,000+ rows)
            col_widths = {
                'A': 15,  # Date
                'B': 12,  # Time
                'C': 15,  # Rashi
                'D': 10,  # Ansha
                'E': 10,  # Kala
                'F': 10,  # Vikala
                'G': 15   # Ayanamsa_DM
            }
            for col_letter, width in col_widths.items():
                worksheet.column_dimensions[col_letter].width = width
                
            # 3. Enable sheet gridlines
            worksheet.views.sheetView[0].showGridLines = True

if __name__ == "__main__":
    app = SunTransitApp()
    app.mainloop()
