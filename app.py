from __future__ import annotations

import calendar
import io
import os
from datetime import date as date_type
from datetime import datetime, timedelta, timezone
from pathlib import Path
from uuid import uuid4

from flask import Flask, abort, jsonify, render_template, request, send_file

from astronomy import get_sunrise, get_sunset, jd_to_zoned_datetime, local_date_anchor_jd
from dainika_muhurta_service import detect_yogas
from location_service import geocode_city, get_timezone_name, search_locations
from pdf_generation_service import generate_pdf_export
from panchang import get_nakshatra, get_tithi, get_vara_from_date
from panchang_service import generate_location_panchang, resolve_location
from range_generation_service import generate_year_range_exports
from request_parsing import (
    parse_panchang_request,
    parse_pdf_generation_request,
    parse_range_generation_request,
    parse_jain_festivals_request,
)

_DAY_CHOGHADIYA_ORDER = ["Udveg", "Char", "Labh", "Amrit", "Kaal", "Shubh", "Rog"]
_NIGHT_CHOGHADIYA_ORDER = ["Shubh", "Amrit", "Char", "Rog", "Kaal", "Labh", "Udveg"]
_CHOGHADIYA_MEANINGS = {
    "Udveg": "Tension", "Amrit": "Nectar", "Rog": "Illness",
    "Labh": "Gain", "Shubh": "Auspicious", "Char": "Movement", "Kaal": "Loss",
}
_CHOGHADIYA_NATURE = {
    "Udveg": "inauspicious", "Amrit": "auspicious", "Rog": "inauspicious",
    "Labh": "auspicious", "Shubh": "auspicious", "Char": "neutral", "Kaal": "inauspicious",
}
# Vara index: Sun=0 ... Sat=6 -> starting index in the day/night Choghadiya order
_DAY_START_IDX = [0, 3, 6, 2, 5, 1, 4]
_NIGHT_START_IDX = [0, 2, 4, 6, 5, 3, 1]

GENERATED_EXPORTS: dict[str, str] = {}

_FILL_HIGHLY_AUSPICIOUS = None
_FILL_AUSPICIOUS = None
_FILL_CAUTION = None
_FILL_AVOID = None


def _get_fills():
    from openpyxl.styles import PatternFill
    return {
        "highly_auspicious": PatternFill("solid", fgColor="C6EFCE"),
        "auspicious": PatternFill("solid", fgColor="DDEBF7"),
        "caution": PatternFill("solid", fgColor="FFEB9C"),
        "avoid": PatternFill("solid", fgColor="FFC7CE"),
        "neutral": PatternFill("solid", fgColor="F2F2F2"),
    }


def _build_muhurta_workbook(summary_rows: list[dict], match_rows: list[dict]):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fills = _get_fills()
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    summary_headers = ["Date", "Vara", "Tithi", "Nakshatra", "Recommendation", "Active Yoga Count"]
    ws_sum.append(summary_headers)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws_sum[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = f"A1:{get_column_letter(len(summary_headers))}1"

    for row in summary_rows:
        ws_sum.append([
            row["date"], row["vara"], row["tithi"], row["nakshatra"],
            row["recommendation"], row["active_yoga_count"],
        ])
        last = ws_sum.max_row
        fill = fills.get(row["recommendation"], fills["neutral"])
        for col in range(1, len(summary_headers) + 1):
            ws_sum.cell(last, col).fill = fill
            ws_sum.cell(last, col).alignment = Alignment(horizontal="center")

    for col, width in zip(range(1, 7), [14, 12, 8, 20, 20, 18]):
        ws_sum.column_dimensions[get_column_letter(col)].width = width

    # ── Matches sheet ──────────────────────────────────────────────────────
    ws_match = wb.create_sheet("Matches")
    match_headers = ["Date", "Yoga Name", "Nature", "Severity", "Trigger", "Meaning", "Day Recommendation"]
    ws_match.append(match_headers)

    for cell in ws_match[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    ws_match.freeze_panes = "A2"
    ws_match.auto_filter.ref = f"A1:{get_column_letter(len(match_headers))}1"

    for row in match_rows:
        ws_match.append([
            row["date"], row["yoga_name"], row["nature"], row["severity"],
            row["trigger_kind"], row["meaning"], row["recommendation"],
        ])
        last = ws_match.max_row
        fill = fills.get(
            "highly_auspicious" if row["nature"] == "shubh" and row["severity"] == "highly_auspicious"
            else "auspicious" if row["nature"] == "shubh"
            else "avoid" if row["severity"] == "highly_inauspicious"
            else "caution"
        )
        for col in range(1, len(match_headers) + 1):
            c = ws_match.cell(last, col)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True)

    for col, width in zip(range(1, 8), [14, 25, 12, 20, 10, 60, 20]):
        ws_match.column_dimensions[get_column_letter(col)].width = width

    return wb


def create_app() -> Flask:
    app = Flask(__name__)

    @app.get("/")
    def index():
        return render_template("index.html")

    @app.get("/search-location")
    def search_location():
        query = request.args.get("q", "").strip()
        if not query:
            return jsonify({"results": []})

        try:
            return jsonify({"results": search_locations(query)})
        except Exception as exc:
            return jsonify({"error": str(exc)}), 502

    @app.get("/get-coordinates")
    def get_coordinates():
        city = request.args.get("city", "").strip()
        if not city:
            return jsonify({"error": "Missing required query parameter: city"}), 400

        try:
            result = geocode_city(city)
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
        except Exception as exc:
            return jsonify({"error": str(exc)}), 502

    @app.post("/generate-panchang")
    def generate_panchang():
        try:
            body = request.get_json(silent=True) or {}
            profile = body.get("profile")
            parsed = parse_panchang_request(body)
            result = generate_location_panchang(
                parsed.input_date,
                city=parsed.city,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa_name=parsed.ayanamsa_name,
            )
            if profile:
                from jain_festival_service import generate_jain_festivals
                d_obj = datetime.strptime(parsed.input_date, "%Y-%m-%d").date()
                fest_data = generate_jain_festivals(d_obj.year, result["lat"], result["lon"], parsed.ayanamsa_name, profile)
                day_festivals = []
                for f in fest_data.get("festivals", []):
                    start_d = datetime.strptime(f["start_date"], "%Y-%m-%d").date()
                    end_d = datetime.strptime(f["end_date"], "%Y-%m-%d").date()
                    if start_d <= d_obj <= end_d:
                        day_festivals.append({
                            "occurrence_id": f["occurrence_id"],
                            "name": f["name"],
                            "category": f["category"],
                            "status": f["status"]
                        })
                result["panchang"]["jain_festivals"] = day_festivals
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/generate-range-panchang")
    def generate_range_panchang():
        try:
            parsed = parse_range_generation_request(request.get_json(silent=True))
            result = generate_year_range_exports(
                start_year=parsed.start_year,
                end_year=parsed.end_year,
                city=parsed.city,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa_name=parsed.ayanamsa_name,
                output_format=parsed.output_format,
                monthly=parsed.monthly,
                workers=parsed.workers,
                profile=parsed.profile,
            )

            files = []
            for generated_file in result["files"]:
                token = uuid4().hex
                GENERATED_EXPORTS[token] = generated_file.path
                files.append(
                    {
                        "name": generated_file.name,
                        "download_url": f"/downloads/{token}",
                    }
                )

            return jsonify(
                {
                    **{k: v for k, v in result.items() if k != "files"},
                    "files": files,
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/generate-pdf-panchang")
    def generate_pdf_panchang():
        try:
            parsed = parse_pdf_generation_request(request.get_json(silent=True))
            result = generate_pdf_export(
                year=parsed.year,
                city=parsed.city,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa_name=parsed.ayanamsa_name,
            )
            token = uuid4().hex
            GENERATED_EXPORTS[token] = result["file"]["path"]

            return jsonify(
                {
                    "year": result["year"],
                    "ayanamsa": result["ayanamsa"],
                    "location": result["location"],
                    "file": {
                        "name": result["file"]["name"],
                        "download_url": f"/downloads/{token}",
                    },
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.get("/month-overview")
    def month_overview():
        try:
            year_str = request.args.get("year")
            month_str = request.args.get("month")

            if not year_str:
                return jsonify({"error": "Missing required parameter: year"}), 400
            if not month_str:
                return jsonify({"error": "Missing required parameter: month"}), 400

            try:
                year = int(year_str)
                month = int(month_str)
            except ValueError:
                return jsonify({"error": "year and month must be integers"}), 400

            if not (1 <= month <= 12):
                return jsonify({"error": "month must be between 1 and 12"}), 400
            if not (1900 <= year <= 2200):
                return jsonify({"error": "year must be between 1900 and 2200"}), 400

            city = request.args.get("city") or None
            lat_str = request.args.get("lat")
            lon_str = request.args.get("lon")
            ayanamsa = request.args.get("ayanamsa", "Lahiri")

            lat = float(lat_str) if lat_str else None
            lon = float(lon_str) if lon_str else None

            if not city and not (lat is not None and lon is not None):
                return jsonify({"error": "Provide either a city name or both latitude and longitude."}), 400

            location = resolve_location(city=city, lat=lat, lon=lon)

            num_days = calendar.monthrange(year, month)[1]
            days = []
            first_result = None

            profile = request.args.get("profile")
            date_to_festivals = {}
            if profile:
                from jain_festival_service import generate_jain_festivals
                fest_data = generate_jain_festivals(year, location.lat, location.lon, ayanamsa, profile)
                for f in fest_data.get("festivals", []):
                    start_d = datetime.strptime(f["start_date"], "%Y-%m-%d").date()
                    end_d = datetime.strptime(f["end_date"], "%Y-%m-%d").date()
                    curr_d = start_d
                    while curr_d <= end_d:
                        if curr_d.year == year and curr_d.month == month:
                            d_str = curr_d.isoformat()
                            if d_str not in date_to_festivals:
                                date_to_festivals[d_str] = []
                            if not any(x["occurrence_id"] == f["occurrence_id"] for x in date_to_festivals[d_str]):
                                date_to_festivals[d_str].append({
                                    "occurrence_id": f["occurrence_id"],
                                    "name": f["name"],
                                    "category": f["category"],
                                    "status": f["status"]
                                })
                        curr_d += timedelta(days=1)

            for day_num in range(1, num_days + 1):
                date_str = f"{year:04d}-{month:02d}-{day_num:02d}"
                result = generate_location_panchang(
                    date_str,
                    lat=location.lat,
                    lon=location.lon,
                    ayanamsa_name=ayanamsa,
                )
                if first_result is None:
                    first_result = result

                udaya_tithi = result["panchang"]["tithi"][0]
                tithi_index = udaya_tithi["index"]
                nakshatra_index = result["panchang"]["nakshatra"]["index"]
                vara_index = result["panchang"]["vara"]["index"]

                tithi_end_raw = udaya_tithi["ends"]["time"] if udaya_tithi["ends"] else ""
                nakshatra_end_raw = result["panchang"]["nakshatra"]["ends"]["time"]
                
                day_payload = {
                    "date": date_str,
                    "tithi_index": tithi_index,
                    "tithi_name": udaya_tithi["name"],
                    "tithi_end_time": tithi_end_raw[:5] if tithi_end_raw else "",
                    "nakshatra_index": nakshatra_index,
                    "nakshatra_name": result["panchang"]["nakshatra"]["name"],
                    "nakshatra_end_time": nakshatra_end_raw[:5] if nakshatra_end_raw else "",
                    "vara_index": vara_index,
                    "vara_name": result["panchang"]["vara"]["name"],
                    "is_purnima": tithi_index == 15,
                    "is_amavasya": tithi_index == 30,
                    "is_ekadashi": tithi_index in (11, 26),
                    "sunrise_time": result["events"]["sunrise"]["time"][:5]
                        if result.get("events", {}).get("sunrise", {}).get("time") else "",
                    "sunset_time": result["events"]["sunset"]["time"][:5]
                        if result.get("events", {}).get("sunset", {}).get("time") else "",
                    "has_panchak": result.get("panchak_kaal", {}).get("has_window", False),
                }
                if profile:
                    day_payload["jain_festivals"] = date_to_festivals.get(date_str, [])
                days.append(day_payload)

            hindu_month_index = first_result["panchang"]["hindu_month"]["index"] if first_result else 0
            hindu_month = first_result["panchang"]["hindu_month"]["name"] if first_result else ""
            vikram_samvat = first_result["panchang"]["vikram_samvat"] if first_result else year + 57

            return jsonify({
                "year": year,
                "month": month,
                "location": location.name,
                "timezone": location.timezone,
                "hindu_month": hindu_month,
                "hindu_month_index": hindu_month_index,
                "vikram_samvat": vikram_samvat,
                "days": days,
            })
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/generate-jain-festivals")
    def generate_jain_festivals_api():
        try:
            body = request.get_json(silent=True) or {}
            parsed = parse_jain_festivals_request(body)
            from jain_festival_service import generate_jain_festivals
            result = generate_jain_festivals(
                year=parsed.year,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa=parsed.ayanamsa_name,
                profile=parsed.profile
            )
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/generate-jain-festival-exports")
    def generate_jain_festival_exports_api():
        try:
            body = request.get_json(silent=True) or {}
            parsed = parse_jain_festivals_request(body)
            fmt = body.get("format", "csv").lower()
            if fmt not in {"csv", "excel", "json", "pdf", "all"}:
                return jsonify({"error": "format must be one of: csv, excel, json, pdf, all"}), 400
                
            from jain_festival_service import generate_jain_festivals
            result = generate_jain_festivals(
                year=parsed.year,
                lat=parsed.lat,
                lon=parsed.lon,
                ayanamsa=parsed.ayanamsa_name,
                profile=parsed.profile
            )
            
            flat_rows = []
            for f in result.get("festivals", []):
                flat_rows.append({
                    "Festival_ID": f["id"],
                    "Name_English": f["name"],
                    "Name_Hindi": f["name_hindi"],
                    "Category": f["category"],
                    "Start_Date": f["start_date"],
                    "End_Date": f["end_date"],
                    "Jain_Month": f["jain_month"],
                    "Paksha": f["paksha"],
                    "Tithi": f["tithi"],
                    "Profile": f["profile"],
                    "Status": f["status"],
                    "Meaning": f["meaning"],
                    "Observance": f["observance"],
                    "Sources": "; ".join(f["sources"])
                })
                
            import tempfile
            import os
            from uuid import uuid4
            
            tmpdir = tempfile.gettempdir()
            unique_id = uuid4().hex
            files = []
            
            formats_to_gen = ["csv", "json", "excel", "pdf"] if fmt == "all" else [fmt]
            for f_format in formats_to_gen:
                if f_format == "csv":
                    filename = f"jain_festivals_{parsed.year}_{unique_id}.csv"
                    path = os.path.join(tmpdir, filename)
                    import csv
                    with open(path, 'w', newline='', encoding='utf-8-sig') as f_out:
                        writer = csv.DictWriter(f_out, fieldnames=list(flat_rows[0].keys()) if flat_rows else [])
                        writer.writeheader()
                        writer.writerows(flat_rows)
                    token = uuid4().hex
                    GENERATED_EXPORTS[token] = path
                    files.append({
                        "name": f"jain_festivals_{parsed.year}.csv",
                        "download_url": f"/downloads/{token}"
                    })
                elif f_format == "json":
                    filename = f"jain_festivals_{parsed.year}_{unique_id}.json"
                    path = os.path.join(tmpdir, filename)
                    import json
                    with open(path, 'w', encoding='utf-8') as f_out:
                        json.dump(flat_rows, f_out, indent=2, ensure_ascii=False)
                    token = uuid4().hex
                    GENERATED_EXPORTS[token] = path
                    files.append({
                        "name": f"jain_festivals_{parsed.year}.json",
                        "download_url": f"/downloads/{token}"
                    })
                elif f_format == "excel":
                    filename = f"jain_festivals_{parsed.year}_{unique_id}.xlsx"
                    path = os.path.join(tmpdir, filename)
                    import pandas as pd
                    df = pd.DataFrame(flat_rows)
                    df.to_excel(path, index=False, engine='openpyxl')
                    token = uuid4().hex
                    GENERATED_EXPORTS[token] = path
                    files.append({
                        "name": f"jain_festivals_{parsed.year}.xlsx",
                        "download_url": f"/downloads/{token}"
                    })
                elif f_format == "pdf":
                    filename = f"jain_festivals_{parsed.year}_{unique_id}.pdf"
                    path = os.path.join(tmpdir, filename)
                    
                    from reportlab.lib.pagesizes import letter
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib import colors
                    
                    doc = SimpleDocTemplate(path, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
                    styles = getSampleStyleSheet()
                    
                    title_style = ParagraphStyle(
                        'TitleStyle',
                        parent=styles['Heading1'],
                        fontSize=18,
                        textColor=colors.HexColor('#2C3E50'),
                        spaceAfter=12
                    )
                    
                    story = []
                    story.append(Paragraph(f"Jain Festival Calendar - {parsed.year}", title_style))
                    story.append(Paragraph(f"Profile: {parsed.profile.replace('_', ' ').title()}", styles['Normal']))
                    story.append(Paragraph(f"Location: Lat {parsed.lat}, Lon {parsed.lon} ({result['location']['timezone']})", styles['Normal']))
                    story.append(Spacer(1, 12))
                    
                    headers = ["Date", "Name (English)", "Category", "Jain Month", "Tithi", "Observance"]
                    table_data = [headers]
                    
                    for row in flat_rows:
                        obs_para = Paragraph(row["Observance"], styles["BodyText"])
                        table_data.append([
                            row["Start_Date"],
                            row["Name_English"],
                            row["Category"].title(),
                            row["Jain_Month"],
                            str(row["Tithi"]),
                            obs_para
                        ])
                        
                    t = Table(table_data, colWidths=[70, 110, 60, 70, 40, 190])
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C3E50')),
                        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('VALIGN', (0,0), (-1,-1), 'TOP'),
                        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0,0), (-1,0), 10),
                        ('BOTTOMPADDING', (0,0), (-1,0), 6),
                        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F8F9FA')),
                        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F4F4')]),
                        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D0D3D4')),
                        ('FONTSIZE', (0,1), (-1,-1), 9),
                    ]))
                    
                    story.append(t)
                    doc.build(story)
                    
                    token = uuid4().hex
                    GENERATED_EXPORTS[token] = path
                    files.append({
                        "name": f"jain_festivals_{parsed.year}.pdf",
                        "download_url": f"/downloads/{token}"
                    })
                    
            return jsonify({
                "year": parsed.year,
                "profile": parsed.profile,
                "files": files
            })
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/choghadiya")
    def choghadiya():
        try:
            body = request.get_json(silent=True) or {}
            date_str = body.get("date")
            lat = body.get("lat")
            lon = body.get("lon")

            if not date_str:
                return jsonify({"error": "Missing required field: date"}), 400
            if lat is None or lon is None:
                return jsonify({"error": "Missing required fields: lat and lon"}), 400

            try:
                parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"error": "date must be in YYYY-MM-DD format"}), 400

            tz_name = get_timezone_name(float(lat), float(lon))
            anchor_jd = local_date_anchor_jd(parsed_date, tz_name)
            sunrise_jd = get_sunrise(anchor_jd, float(lat), float(lon))
            sunset_jd = get_sunset(anchor_jd, float(lat), float(lon))

            next_date = parsed_date + timedelta(days=1)
            next_anchor_jd = local_date_anchor_jd(next_date, tz_name)
            next_sunrise_jd = get_sunrise(next_anchor_jd, float(lat), float(lon))

            if not sunrise_jd or not sunset_jd or not next_sunrise_jd:
                return jsonify({"error": "Sunrise or sunset could not be calculated for this location/date."}), 400

            weekday = (parsed_date.weekday() + 1) % 7
            day_start = _DAY_START_IDX[weekday]
            night_start = _NIGHT_START_IDX[weekday]

            def _make_slots(
                start_jd: float,
                end_jd: float,
                choghadiya_order: list[str],
                start_idx: int,
                period: str,
            ) -> list[dict]:
                slot_duration = (end_jd - start_jd) / 8
                slot_duration_minutes = slot_duration * 1440
                slots = []
                for i in range(8):
                    name = choghadiya_order[(start_idx + i) % 7]
                    slot_start = start_jd + i * slot_duration
                    slot_end = start_jd + (i + 1) * slot_duration
                    start_dt = jd_to_zoned_datetime(slot_start, tz_name)
                    end_dt = jd_to_zoned_datetime(slot_end, tz_name)
                    start_utc = start_dt.astimezone(timezone.utc) if start_dt else None
                    end_utc = end_dt.astimezone(timezone.utc) if end_dt else None

                    def _label(local_dt: datetime | None) -> str:
                        if local_dt is None:
                            return ""
                        time_label = local_dt.strftime("%I:%M %p").lstrip("0")
                        if local_dt.date() == parsed_date:
                            return time_label
                        return f"{time_label}, {local_dt.strftime('%B')} {local_dt.day}"

                    slots.append({
                        "name": name,
                        "meaning": _CHOGHADIYA_MEANINGS[name],
                        "nature": _CHOGHADIYA_NATURE[name],
                        "start_time": start_dt.strftime("%H:%M") if start_dt else "",
                        "end_time": end_dt.strftime("%H:%M") if end_dt else "",
                        "start_local": start_dt.isoformat(timespec="seconds") if start_dt else "",
                        "end_local": end_dt.isoformat(timespec="seconds") if end_dt else "",
                        "start_utc": start_utc.isoformat(timespec="seconds") if start_utc else "",
                        "end_utc": end_utc.isoformat(timespec="seconds") if end_utc else "",
                        "start_label": _label(start_dt),
                        "end_label": _label(end_dt),
                        "duration_minutes": round(slot_duration_minutes, 3),
                        "period": period,
                    })
                return slots

            sunrise_dt = jd_to_zoned_datetime(sunrise_jd, tz_name)
            sunset_dt = jd_to_zoned_datetime(sunset_jd, tz_name)
            sunrise_utc = sunrise_dt.astimezone(timezone.utc) if sunrise_dt else None
            sunset_utc = sunset_dt.astimezone(timezone.utc) if sunset_dt else None
            day_slot_duration_minutes = (sunset_jd - sunrise_jd) * 1440 / 8
            night_slot_duration_minutes = (next_sunrise_jd - sunset_jd) * 1440 / 8

            slots = (
                _make_slots(sunrise_jd, sunset_jd, _DAY_CHOGHADIYA_ORDER, day_start, "day")
                + _make_slots(sunset_jd, next_sunrise_jd, _NIGHT_CHOGHADIYA_ORDER, night_start, "night")
            )

            return jsonify({
                "date": date_str,
                "timezone": tz_name,
                "sunrise": sunrise_dt.strftime("%H:%M") if sunrise_dt else "",
                "sunset": sunset_dt.strftime("%H:%M") if sunset_dt else "",
                "sunrise_local": sunrise_dt.isoformat(timespec="seconds") if sunrise_dt else "",
                "sunset_local": sunset_dt.isoformat(timespec="seconds") if sunset_dt else "",
                "sunrise_utc": sunrise_utc.isoformat(timespec="seconds") if sunrise_utc else "",
                "sunset_utc": sunset_utc.isoformat(timespec="seconds") if sunset_utc else "",
                "day_slot_duration_minutes": round(day_slot_duration_minutes, 3),
                "night_slot_duration_minutes": round(night_slot_duration_minutes, 3),
                "slots": slots,
            })
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/dainika-muhurta")
    def dainika_muhurta():
        try:
            body = request.get_json(silent=True) or {}
            date_str = body.get("date")
            lat = body.get("lat")
            lon = body.get("lon")
            ayanamsa = body.get("ayanamsa", "Lahiri")

            if not date_str:
                return jsonify({"error": "Missing required field: date"}), 400
            if lat is None or lon is None:
                return jsonify({"error": "Missing required fields: lat and lon"}), 400

            try:
                parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"error": "date must be in YYYY-MM-DD format"}), 400

            tz_name = get_timezone_name(float(lat), float(lon))
            anchor_jd = local_date_anchor_jd(parsed_date, tz_name)
            sunrise_jd = get_sunrise(anchor_jd, float(lat), float(lon))

            if not sunrise_jd:
                return jsonify({"error": "Sunrise could not be calculated for this location/date."}), 400

            from astronomy import get_ayanamsa, get_planetary_longitude
            ayanamsa_val = get_ayanamsa(sunrise_jd, ayanamsa)
            sun_lon = (get_planetary_longitude(sunrise_jd, "Sun") - ayanamsa_val) % 360.0
            moon_lon = (get_planetary_longitude(sunrise_jd, "Moon") - ayanamsa_val) % 360.0

            vara = get_vara_from_date(parsed_date)
            tithi = get_tithi(sun_lon, moon_lon)
            nakshatra = get_nakshatra(moon_lon)

            yoga_result = detect_yogas(vara=vara, tithi=tithi, nakshatra=nakshatra)

            return jsonify({
                "date": date_str,
                "vara": vara,
                "tithi": tithi,
                "nakshatra": nakshatra,
                "yogas": yoga_result["yogas"],
                "recommendation": yoga_result["recommendation"],
            })
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.post("/dainika-muhurta-export")
    def dainika_muhurta_export():
        try:
            body = request.get_json(silent=True) or {}
            year = body.get("year")
            month = body.get("month")
            lat = body.get("lat")
            lon = body.get("lon")
            ayanamsa = body.get("ayanamsa", "Lahiri")

            if year is None:
                return jsonify({"error": "Missing required field: year"}), 400
            if month is None:
                return jsonify({"error": "Missing required field: month"}), 400
            if lat is None or lon is None:
                return jsonify({"error": "Missing required fields: lat and lon"}), 400

            year, month = int(year), int(month)
            if not (1 <= month <= 12):
                return jsonify({"error": "month must be 1–12"}), 400

            tz_name = get_timezone_name(float(lat), float(lon))
            days_in_month = calendar.monthrange(year, month)[1]

            from astronomy import get_ayanamsa, get_planetary_longitude

            summary_rows = []
            match_rows = []

            NAKSHATRA_NAMES = [
                "", "Ashvini", "Bharani", "Kritika", "Rohini", "Mrigashira",
                "Ardra", "Punarvasu", "Pushya", "Ashlesha", "Magha",
                "Purva Phalguni", "Uttara Phalguni", "Hasta", "Chitra", "Swati",
                "Vishakha", "Anuradha", "Jyeshtha", "Mula", "Purva Ashadha",
                "Uttara Ashadha", "Shravana", "Dhanishtha", "Shatabhisha",
                "Purva Bhadrapada", "Uttara Bhadrapada", "Revati", "Abhijit",
            ]
            VARA_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

            for day in range(1, days_in_month + 1):
                d = date_type(year, month, day)
                anchor_jd = local_date_anchor_jd(d, tz_name)
                sunrise_jd = get_sunrise(anchor_jd, float(lat), float(lon))
                if not sunrise_jd:
                    continue

                ayanamsa_val = get_ayanamsa(sunrise_jd, ayanamsa)
                sun_lon = (get_planetary_longitude(sunrise_jd, "Sun") - ayanamsa_val) % 360.0
                moon_lon = (get_planetary_longitude(sunrise_jd, "Moon") - ayanamsa_val) % 360.0

                vara = get_vara_from_date(d)
                tithi = get_tithi(sun_lon, moon_lon)
                nakshatra = get_nakshatra(moon_lon)
                yoga_result = detect_yogas(vara=vara, tithi=tithi, nakshatra=nakshatra)

                summary_rows.append({
                    "date": d.strftime("%Y-%m-%d"),
                    "vara": VARA_NAMES[vara],
                    "tithi": tithi,
                    "nakshatra": NAKSHATRA_NAMES[nakshatra] if nakshatra < len(NAKSHATRA_NAMES) else str(nakshatra),
                    "recommendation": yoga_result["recommendation"],
                    "active_yoga_count": len([y for y in yoga_result["yogas"] if not y.get("cancelled")]),
                })
                for yoga in yoga_result["yogas"]:
                    if yoga.get("cancelled"):
                        continue
                    match_rows.append({
                        "date": d.strftime("%Y-%m-%d"),
                        "yoga_name": yoga["name"],
                        "nature": yoga["nature"],
                        "severity": yoga["severity"],
                        "trigger_kind": yoga["trigger_kind"],
                        "meaning": yoga["meaning"],
                        "recommendation": yoga_result["recommendation"],
                    })

            wb = _build_muhurta_workbook(summary_rows, match_rows)
            filename = f"dainika_muhurta_{year}_{str(month).zfill(2)}.xlsx"

            tmp_dir = Path("output")
            tmp_dir.mkdir(exist_ok=True)
            file_path = tmp_dir / filename
            wb.save(str(file_path))

            token = str(uuid4())
            GENERATED_EXPORTS[token] = str(file_path)

            return jsonify({
                "filename": filename,
                "download_url": f"/downloads/{token}",
            })
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @app.get("/downloads/<token>")
    def download_generated_file(token: str):
        file_path = GENERATED_EXPORTS.get(token)
        if not file_path:
            abort(404)

        path = Path(file_path)
        if not path.exists():
            abort(404)

        return send_file(path, as_attachment=True, download_name=path.name)

    return app


app = create_app()


if __name__ == "__main__":
    app.run(debug=True)
