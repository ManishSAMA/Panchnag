import os
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def save_to_excel(rows: list[dict], file_path: str, year: int):
    """Write rows to Excel using pandas/openpyxl and style the header and column widths."""
    df = pd.DataFrame(rows)
    sheet_name = f"Sun Transits {year}"
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # 1. Colors & Fills for Headers (Professional deep blue)
        header_fill = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Style header row (Row 1)
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
        # 2. Predefined Column dimensions
        col_widths = {
            'A': 15,  # Date
            'B': 12,  # Time
            'C': 15,  # Rashi
            'D': 10,  # Ansha
            'E': 10,  # Kala
            'F': 10,  # Vikala
            'G': 15   # Ayanamsa_DM
        }
        for col_letter, width in col_widths.items():
            worksheet.column_dimensions[col_letter].width = width
            
        # 3. Enable sheet gridlines
        worksheet.views.sheetView[0].showGridLines = True
