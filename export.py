"""
export.py — Data formatting and file export utilities for the Panchang Generator.

Supports export to:
  - CSV   (comma-separated values, UTF-8)
  - Excel (.xlsx via openpyxl)
  - JSON  (pretty-printed)

Each row stores both raw decimal-degree values (machine-readable)
and human-readable DMS strings for every planetary longitude.
"""

import csv
import json
import os

from astronomy import format_dms, get_sun_rashi, jd_to_local_datetime_string, jd_to_local_time_string
from panchang import get_hindu_month

ELEMENT_DISPLAY_SPECS = (
    ("Tithi", "Tithi_End_JD"),
    ("Nakshatra", "Nakshatra_End_JD"),
    ("Yoga", "Yoga_End_JD"),
)

# ---------------------------------------------------------------------------
# Row Formatter
# ---------------------------------------------------------------------------

def format_row_data(
    date_str: str,
    julian_date: float,
    planets: dict[str, float],   # name → decimal degrees
    panchang: dict,
    jain_tithi: dict,
    sunrise_str: str,
    sunset_str: str,
    moonrise_str: str,
    moonset_str: str,
    ayanamsa_dec: float,
    ayanamsa_name: str = 'Lahiri',
    tz_offset: float = 5.5,
    tz_label: str = "IST",
    vikram_samvat: int | None = None,
    vira_nirvana_samvat: int | None = None,
) -> dict:
    """Build a flat dict representing one row of the Panchang table.

    Args:
        date_str:      ISO date string (YYYY-MM-DD).
        julian_date:   Julian Day Number.
        planets:       Dict of planet→decimal longitude (sidereal).
        panchang:      Output of generate_daily_panchang().
        sunrise_str:   Formatted local time string for sunrise.
        sunset_str:    Formatted local time string for sunset.
        moonrise_str:  Formatted local time string for moonrise.
        moonset_str:   Formatted local time string for moonset.
        ayanamsa_dec:  Ayanamsa value in decimal degrees.
        tz_label:      Timezone label used in column headers.

    Returns:
        Ordered flat dict suitable for CSV / Excel / JSON export.
    """
    row: dict = {}

    # ---- Date & JD ----
    row['Date'] = date_str
    if vikram_samvat is not None:
        row['Vikram_Samvat'] = vikram_samvat
    if vira_nirvana_samvat is not None:
        row['Vira_Nirvana_Samvat'] = vira_nirvana_samvat
    row['Julian_Date'] = round(julian_date, 5)
    row['Weekday'] = panchang['Vara_Name']
    row['Sun_Rashi'] = get_sun_rashi(julian_date)

    # ---- Hindu Lunar Month ----
    hindu_month, hindu_month_common, _is_adhika = get_hindu_month(julian_date, ayanamsa_name)
    row['Hindu_Month'] = hindu_month
    row['Hindu_Month_Common'] = hindu_month_common

    # ---- Panchang Elements ----
    row['Tithi_No']    = panchang['Tithi_Index']
    row['Tithi']       = panchang['Tithi_Name']
    row['__Tithi_End_JD'] = panchang.get('Tithi_End_JD', 0.0)
    row['Jain_Tithi_No'] = jain_tithi['Jain_Tithi_Index']
    row['Jain_Tithi'] = jain_tithi['Jain_Tithi_Name']
    row['Jain_Tithi_End_Time'] = jd_to_local_time_string(jain_tithi['Jain_Tithi_End_JD'], tz_offset)
    row['Nakshatra_No']= panchang['Nakshatra_Index']
    row['Nakshatra']   = panchang['Nakshatra_Name']
    row['Nakshatra_Pada'] = panchang['Nakshatra_Pada']
    row['__Nakshatra_End_JD'] = panchang.get('Nakshatra_End_JD', 0.0)
    row['Yoga_No']     = panchang['Yoga_Index']
    row['Yoga']        = panchang['Yoga_Name']
    row['__Yoga_End_JD'] = panchang.get('Yoga_End_JD', 0.0)
    row['Karana_No']   = panchang['Karana_Index']
    row['Karana']      = panchang['Karana_Name']
    row['Karana Start'] = jd_to_local_time_string(panchang['Karana_Start_JD'], tz_offset)[:5]
    row['Karana End']   = jd_to_local_time_string(panchang['Karana_End_JD'], tz_offset)[:5]

    # ---- Sunrise / Sunset / Moonrise / Moonset ----
    row[f'Sunrise ({tz_label})']  = sunrise_str
    row[f'Sunset ({tz_label})']   = sunset_str
    row[f'Moonrise ({tz_label})'] = moonrise_str
    row[f'Moonset ({tz_label})']  = moonset_str

    # ---- Planetary Longitudes (decimal & DMS) ----
    planet_order = ['Sun', 'Moon', 'Mars', 'Mercury', 'Jupiter',
                    'Venus', 'Saturn', 'Rahu', 'Ketu']
    for planet in planet_order:
        dec = planets.get(planet, 0.0)
        row[f'{planet}_Dec']  = round(dec, 6)
        row[f'{planet}_DMS']  = format_dms(dec)

    # ---- Ayanamsa ----
    row['Ayanamsa_Dec'] = round(ayanamsa_dec, 6)
    row['Ayanamsa_DMS'] = format_dms(ayanamsa_dec)

    return row


def _format_element_value(name: str, end_jd: float, show_end_info: bool, tz_offset: float) -> str:
    if not name:
        return ""
    if not show_end_info or not end_jd:
        return name

    end_label = jd_to_local_datetime_string(end_jd, tz_offset)
    if not end_label:
        return name
    return f"{name} [{end_label}]"


def apply_element_continuity_formatting(data_list: list[dict], tz_offset: float = 5.5) -> list[dict]:
    """Apply shared continuity formatting for repeating Panchang elements."""
    if not data_list:
        return []

    formatted_rows: list[dict] = []
    for index, source_row in enumerate(data_list):
        row = dict(source_row)
        for element_key, end_key in ELEMENT_DISPLAY_SPECS:
            current_name = source_row.get(element_key, "")
            prev_name = data_list[index - 1].get(element_key, "") if index > 0 else ""
            next_name = data_list[index + 1].get(element_key, "") if index + 1 < len(data_list) else ""

            is_first_occurrence = current_name != prev_name
            is_continuation = current_name == prev_name
            is_last_occurrence = current_name != next_name

            row[f"__{element_key}_Is_First_Occurrence"] = is_first_occurrence
            row[f"__{element_key}_Is_Continuation"] = is_continuation
            row[f"__{element_key}_Is_Last_Occurrence"] = is_last_occurrence
            row[element_key] = _format_element_value(
                current_name,
                float(source_row.get(f"__{end_key}", 0.0) or 0.0),
                show_end_info=is_last_occurrence,
                tz_offset=tz_offset,
            )

        row["Jain_Tithi_PDF"] = source_row.get("Jain_Tithi", "")
        formatted_rows.append(row)

    return formatted_rows


def strip_internal_output_fields(data_list: list[dict]) -> list[dict]:
    """Remove helper fields before writing rows to flat file outputs."""
    cleaned_rows: list[dict] = []
    for row in data_list:
        cleaned_rows.append({
            key: value
            for key, value in row.items()
            if not key.startswith("__") and key != "Jain_Tithi_PDF"
        })
    return cleaned_rows


# ---------------------------------------------------------------------------
# Export Functions
# ---------------------------------------------------------------------------

def export_to_csv(data_list: list[dict], filename: str = "panchang_output.csv") -> None:
    """Export data to a UTF-8 CSV file."""
    if not data_list:
        print("No data to export.")
        return
    clean_rows = strip_internal_output_fields(data_list)
    keys = list(clean_rows[0].keys())
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(clean_rows)
    print(f"  ✓ CSV  → {os.path.abspath(filename)}")


def export_to_json(data_list: list[dict], filename: str = "panchang_output.json") -> None:
    """Export data to a pretty-printed JSON file."""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(strip_internal_output_fields(data_list), f, indent=2, ensure_ascii=False)
    print(f"  ✓ JSON → {os.path.abspath(filename)}")


def export_to_excel(data_list: list[dict], filename: str = "panchang_output.xlsx") -> None:
    """Export data to an Excel workbook (.xlsx)."""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("pandas not installed. Run: pip install pandas openpyxl")
        return
    if not data_list:
        print("No data to export.")
        return
    df = pd.DataFrame(strip_internal_output_fields(data_list))
    df.to_excel(filename, index=False, engine='openpyxl')

    workbook = load_workbook(filename)
    worksheet = workbook.active
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(fill_type="solid", fgColor="2C3E50")
    accent_fill = PatternFill(fill_type="solid", fgColor="E8F1FB")
    thin_side = Side(style="thin", color="D0D7DE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    centered = Alignment(horizontal="center", vertical="center")

    accent_headers = {"Tithi_No", "Tithi", "Jain_Tithi_No", "Jain_Tithi", "Jain_Tithi_End_Time"}

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = border

    header_map = {cell.column: cell.value for cell in worksheet[1]}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = centered
            cell.border = border
            if header_map.get(cell.column) in accent_headers:
                cell.fill = accent_fill

    for index, column_name in enumerate(df.columns, start=1):
        max_length = max(len(str(column_name)), *(len(str(value)) for value in df[column_name].fillna("")))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 28)

    workbook.save(filename)
    print(f"  ✓ Excel→ {os.path.abspath(filename)}")


def export_data(data_list: list[dict], base_filename: str, fmt: str) -> None:
    """Dispatch export to the correct format.

    Args:
        data_list:      List of row dicts.
        base_filename:  Base path without extension.
        fmt:            One of 'csv', 'excel', 'json', 'all'.
    """
    if not data_list:
        print("Warning: No data rows to export.")
        return

    formats = ['csv', 'excel', 'json'] if fmt == 'all' else [fmt]
    for f in formats:
        ext_map = {'csv': 'csv', 'excel': 'xlsx', 'json': 'json'}
        fname = f"{base_filename}.{ext_map[f]}"
        if f == 'csv':
            export_to_csv(data_list, fname)
        elif f == 'json':
            export_to_json(data_list, fname)
        elif f == 'excel':
            export_to_excel(data_list, fname)
