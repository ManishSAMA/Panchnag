"""API tests for POST /dainika-muhurta and POST /dainika-muhurta-export."""

import io
import unittest

from app import app


class TestDainikamuhurtaEndpoint(unittest.TestCase):
    def setUp(self):
        self.client = app.test_client()

    def test_valid_request_returns_200(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lat": 26.9124, "lon": 75.7873},
        )
        self.assertEqual(resp.status_code, 200)

    def test_response_has_required_keys(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        for key in ("date", "vara", "tithi", "nakshatra", "yogas", "recommendation"):
            self.assertIn(key, data, f"Missing key: {key}")

    def test_vara_is_integer_0_to_6(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        self.assertIsInstance(data["vara"], int)
        self.assertIn(data["vara"], range(7))

    def test_tithi_is_integer_1_to_30(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        self.assertIsInstance(data["tithi"], int)
        self.assertIn(data["tithi"], range(1, 31))

    def test_nakshatra_is_integer_1_to_27(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        self.assertIsInstance(data["nakshatra"], int)
        self.assertIn(data["nakshatra"], range(1, 29))  # 1-27 + possibly 28 Abhijit

    def test_recommendation_is_valid_value(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        self.assertIn(
            data["recommendation"],
            ("highly_auspicious", "auspicious", "caution", "avoid", "neutral"),
        )

    def test_yogas_is_a_list(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        self.assertIsInstance(data["yogas"], list)

    def test_yoga_entry_has_required_fields(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-01-07", "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        if data["yogas"]:
            yoga = data["yogas"][0]
            for field in ("name", "nature", "trigger_kind", "severity", "meaning"):
                self.assertIn(field, yoga)

    def test_missing_date_returns_400(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"lat": 26.9124, "lon": 75.7873},
        )
        self.assertEqual(resp.status_code, 400)

    def test_missing_lat_returns_400(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lon": 75.7873},
        )
        self.assertEqual(resp.status_code, 400)

    def test_missing_lon_returns_400(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lat": 26.9124},
        )
        self.assertEqual(resp.status_code, 400)

    def test_invalid_date_format_returns_400(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "03-06-2026", "lat": 26.9124, "lon": 75.7873},
        )
        self.assertEqual(resp.status_code, 400)

    def test_ayanamsa_parameter_accepted(self):
        resp = self.client.post(
            "/dainika-muhurta",
            json={"date": "2026-06-03", "lat": 26.9124, "lon": 75.7873, "ayanamsa": "Lahiri"},
        )
        self.assertEqual(resp.status_code, 200)


class TestDainikaExportEndpoint(unittest.TestCase):
    def setUp(self):
        self.client = app.test_client()

    def test_valid_request_returns_200(self):
        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "month": 6, "lat": 26.9124, "lon": 75.7873},
        )
        self.assertEqual(resp.status_code, 200)

    def test_response_has_download_url(self):
        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "month": 6, "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        self.assertIn("download_url", data)

    def test_download_url_token_serves_xlsx(self):
        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "month": 6, "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        token = data["download_url"].split("/")[-1]
        dl = self.client.get(f"/downloads/{token}")
        self.assertEqual(dl.status_code, 200)
        self.assertIn("spreadsheet", dl.content_type)

    def test_missing_year_returns_400(self):
        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"month": 6, "lat": 26.9124, "lon": 75.7873},
        )
        self.assertEqual(resp.status_code, 400)

    def test_missing_month_returns_400(self):
        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "lat": 26.9124, "lon": 75.7873},
        )
        self.assertEqual(resp.status_code, 400)

    def test_missing_coords_returns_400(self):
        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "month": 6},
        )
        self.assertEqual(resp.status_code, 400)

    def test_workbook_has_summary_and_matches_sheets(self):
        import openpyxl

        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "month": 6, "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        token = data["download_url"].split("/")[-1]
        dl = self.client.get(f"/downloads/{token}")
        wb = openpyxl.load_workbook(io.BytesIO(dl.data))
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Matches", wb.sheetnames)

    def test_summary_sheet_has_30_data_rows_for_june(self):
        import openpyxl

        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "month": 6, "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        token = data["download_url"].split("/")[-1]
        dl = self.client.get(f"/downloads/{token}")
        wb = openpyxl.load_workbook(io.BytesIO(dl.data))
        ws = wb["Summary"]
        # Row 1 = header; rows 2..31 = 30 days of June
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        self.assertEqual(len(data_rows), 30)

    def test_summary_header_has_expected_columns(self):
        import openpyxl

        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "month": 6, "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        token = data["download_url"].split("/")[-1]
        dl = self.client.get(f"/downloads/{token}")
        wb = openpyxl.load_workbook(io.BytesIO(dl.data))
        ws = wb["Summary"]
        headers = [c.value for c in ws[1]]
        for col in ("Date", "Vara", "Tithi", "Nakshatra", "Recommendation", "Active Yoga Count"):
            self.assertIn(col, headers)

    def test_matches_sheet_one_row_per_yoga_per_day(self):
        import openpyxl

        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "month": 6, "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        token = data["download_url"].split("/")[-1]
        dl = self.client.get(f"/downloads/{token}")
        wb = openpyxl.load_workbook(io.BytesIO(dl.data))
        ws = wb["Matches"]
        # Row 1 = header. Count data rows.
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        # Each data row must have a Date and a Yoga Name
        for row in data_rows:
            self.assertIsNotNone(row[0])  # Date
            self.assertIsNotNone(row[1])  # Yoga Name

    def test_filename_contains_year_and_month(self):
        resp = self.client.post(
            "/dainika-muhurta-export",
            json={"year": 2026, "month": 6, "lat": 26.9124, "lon": 75.7873},
        )
        data = resp.get_json()
        self.assertIn("filename", data)
        self.assertIn("2026", data["filename"])
        self.assertIn("06", data["filename"])
